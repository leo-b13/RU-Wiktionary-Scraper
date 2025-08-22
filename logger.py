from datetime import datetime, timedelta
import openpyxl
import os 
from openpyxl import Workbook, load_workbook

def load2excel(userInput):
    foundWord = False
    fileName = "russian_words_log.xlsx"
    today = datetime.today().strftime('%Y-%m-%d')
    #yesterday = (datetime.now() - timedelta(1)).strftime('%Y-%m-%d')

    if not os.path.exists(fileName):
        print("Excel file not found, making a new one!")
        wb = Workbook()
        ws = wb.active
        ws.append(["First Searched", "Last Searched", "Word", "Count"])
        wb.save(fileName)

    wb = load_workbook(fileName)
    ws = wb.active

    for row in ws.iter_rows(min_col=3, max_col=3):
        cell = row[0]
        if cell.value == userInput:
            increment_cell = ws.cell(row=cell.row, column=4)
            increment_cell.value = increment_cell.value + 1
            word_counter = increment_cell.value

            date_cell = ws.cell(row=cell.row, column=2)
            word_date = date_cell.value
            date_cell.value = today
            
            foundWord = True
            print("Word found!")
            break

    if not foundWord:
        print("!!! Word not found, making new row!")
        new_row = ws.max_row + 1 
        ws.cell(row=new_row, column=1).value = today
        ws.cell(row=new_row, column=2).value = today
        ws.cell(row=new_row, column=3).value = userInput
        ws.cell(row=new_row, column=4).value = 1
        word_date = today
        word_counter = "New word!"

    wb.save(fileName)

    return word_date, word_counter

def get_excel_data():
    fileName = "russian_words_log.xlsx"
    wb = load_workbook(fileName)
    ws = wb.active

    excel_data = list(ws.values)
    max_r = ws.max_row
    max_c = ws.max_column
    
    return excel_data, max_r, max_c